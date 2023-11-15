

import os
from threading import Thread
import cv2
from django.conf import settings
from django.shortcuts import redirect
from openpyxl import Workbook

from itp_forms.core.config import Config
from itp_forms.core.image_handler import ImageHandler

XLSX_ANSWERS_FOLDER = os.path.join(settings.MEDIA_ROOT, '07_xlsx_answers_folder')

ALPHABET = ('A','B','C','D','E','F','G','H','I','J','K','L','M','N','O','P','Q','R',
    'S','T','U','V','W','X','Y','Z','AA','AB','AC','AD','AE','AF','AG','AH','AI',
    'AJ','AK','AL','AM','AN','AO','AP','AQ','AR','AS','AT','AU','AV','AW','AX','AY',
    'AZ','BA','BB','BC','BD','BE','BF','BG','BH','BI','BJ','BK','BL','BM','BN','BO',
    'BP','BQ','BR','BS','BT','BU','BV','BW','BX','BY','BZ','CA','CB','CC','CD','CE',
    'CF','CG','CH','CI','CJ','CK','CL','CM','CN','CO','CP','CQ','CR','CS','CT','CU',
    'CV','CW','CX','CY','CZ','DA','DB','DC','DD','DE','DF','DG','DH','DI','DJ','DK',
    'DL','DM','DN','DO','DP','DQ','DR','DS','DT','DU','DV','DW','DX','DY','DZ','EA',
    'EB','EC','ED','EE','EF','EG','EH','EI','EJ','EK','EL','EM','EN','EO','EP','EQ',
    'ER','ES','ET','EU','EV','EW','EX','EY','EZ','FA','FB','FC','FD','FE','FF','FG',
    'FH','FI','FJ','FK','FL','FM','FN','FO','FP','FQ','FR','FS','FT','FU','FV','FW',
    'FX','FY','FZ','GA','GB','GC','GD','GE','GF','GG','GH','GI','GJ','GK','GL','GM',
    'GN','GO','GP','GQ','GR','GS','GT','GU','GV','GW','GX','GY','GZ','HA','HB','HC',
    'HD','HE','HF','HG','HH','HI','HJ','HK','HL','HM','HN','HO','HP','HQ','HR','HS',
    'HT','HU','HV','HW','HX','HY','HZ','IA','IB','IC','ID','IE','IF','IG','IH','II',
    'IJ','IK','IL','IM','IN','IO','IP','IQ','IR','IS','IT','IU','IV','IW','IX','IY',
    'IZ','JA','JB','JC','JD','JE','JF','JG','JH','JI','JJ','JK','JL','JM','JN','JO',
    'JP','JQ','JR','JS','JT','JU','JV','JW','JX','JY','JZ','KA','KB','KC','KD','KE',
    'KF','KG','KH','KI','KJ','KK','KL','KM','KN','KO','KP','KQ','KR','KS','KT','KU',
    'KV','KW','KX','KY','KZ','LA','LB','LC','LD','LE','LF','LG','LH','LI','LJ','LK',
    'LL','LM','LN','LO','LP','LQ','LR','LS','LT','LU','LV','LW','LX','LY','LZ','MA',
    'MB','MC','MD','ME','MF','MG','MH','MI','MJ','MK','ML','MM','MN','MO','MP','MQ',
    'MR','MS','MT','MU','MV','MW','MX','MY','MZ','NA','NB','NC','ND','NE','NF','NG',
    'NH','NI','NJ','NK','NL','NM','NN','NO','NP','NQ','NR','NS','NT','NU','NV','NW',
    'NX','NY','NZ','OA','OB','OC','OD','OE','OF','OG','OH','OI','OJ','OK','OL','OM',
    'ON','OO','OP','OQ','OR','OS','OT','OU','OV','OW','OX','OY','OZ','PA','PB','PC',
    'PD','PE','PF','PG','PH','PI','PJ','PK','PL','PM','PN','PO','PP','PQ','PR','PS',
    'PT','PU','PV','PW','PX','PY','PZ','QA','QB','QC','QD','QE','QF','QG','QH','QI',
    'QJ','QK','QL','QM','QN','QO','QP','QQ','QR','QS','QT','QU','QV','QW','QX','QY',
    'QZ','RA','RB','RC','RD','RE','RF','RG','RH','RI','RJ','RK','RL','RM','RN','RO',
    'RP','RQ','RR','RS','RT','RU','RV','RW','RX','RY','RZ','SA','SB','SC','SD','SE',
    'SF','SG','SH','SI','SJ','SK','SL','SM','SN','SO','SP','SQ','SR','SS','ST','SU',
    'SV','SW','SX','SY','SZ','TA','TB','TC','TD','TE','TF','TG','TH','TI','TJ','TK',
    'TL','TM','TN','TO','TP','TQ','TR','TS','TT','TU','TV','TW','TX','TY','TZ','UA',
    'UB','UC','UD','UE','UF','UG','UH','UI','UJ','UK','UL','UM','UN','UO','UP','UQ',
    'UR','US','UT','UU','UV','UW','UX','UY','UZ','VA','VB','VC','VD','VE','VF','VG',
    'VH','VI','VJ','VK','VL','VM','VN','VO','VP','VQ','VR','VS','VT','VU','VV','VW',
    'VX','VY','VZ','WA','WB','WC','WD','WE','WF','WG','WH','WI','WJ','WK','WL','WM',
    'WN','WO','WP','WQ','WR','WS','WT','WU','WV','WW','WX','WY','WZ','XA','XB','XC',
    'XD','XE','XF','XG','XH','XI','XJ','XK','XL','XM','XN','XO','XP','XQ','XR','XS',
    'XT','XU','XV','XW','XX','XY','XZ','YA','YB','YC','YD','YE','YF','YG','YH','YI',
    'YJ','YK','YL','YM','YN','YO','YP','YQ','YR','YS','YT','YU','YV','YW','YX','YY',)
    
class AnswersInterpreter:


    def __init__(self,answers_folder) -> None:
        if not os.path.exists(XLSX_ANSWERS_FOLDER):
            os.makedirs(XLSX_ANSWERS_FOLDER)
        self.answers_folder = answers_folder
        self.wb = Workbook()
        self.ws = self.wb.active
        self.ws.title = f"Exportação de do formulário {os.path.basename(self.answers_folder)}"
        self.set_initial_ws_layout()
        self.information = {}
    
    def set_initial_ws_layout(self):
        self.ws["A1"] = "Questão"
        for row, page in enumerate(os.listdir(self.answers_folder), 1):
            self.ws[f"A{row+1}"] = f"Formulário {row}"
        for question in range(Config.instance().get_questions_amount()):
            self.ws[f"{ALPHABET[question+1]}1"] = question+1

    def interpret_answers(self):
        threads = []
        for row, page in enumerate(os.listdir(self.answers_folder), 1):
            file = os.path.join(self.answers_folder, page) 
            thread = Thread(target=self.interpret_page, args=(file, row))
            threads.append(thread)
            thread.start()

        for thread in threads:
            thread.join()

        results_path = os.path.join(XLSX_ANSWERS_FOLDER, f'{os.path.basename(self.answers_folder)}.xlsx')
        self.wb.save(results_path)
        Config.instance().results_path = results_path
        Config.instance().question_results = self.information
        return redirect("results_view")
        
    def interpret_page(self, file, ws_row_index):     
        
        handler = ImageHandler(base_image_path=file)
        image = handler.cropp_image()

        image_height, image_width, _ = image.shape

        config = Config.instance()
        x_axis_reason, y_axis_reason = config.get_multiply_reason(image_height, image_width)

        x_space_between_cells = config.x_space_between_cells * x_axis_reason
        y_space_between_cells = config.y_space_between_cells * y_axis_reason

        cell_size_x_px = config.cell_size_x_px * x_axis_reason
        cell_size_y_px = config.cell_size_y_px * y_axis_reason
        
        groupings = []

        highest_mean_color = 0
        lowest_mean_color = 999

        # para cada agrupamento
        for i in range(1,5):
            grouping = []

            # verifica se o agrupamento tem todas informações necessarias
            if not config.grouping_has_all_info(i):
                break

            # cada coluna é uma questão
            for row in range(getattr(config, f'grouping_{i}_row_amount')):
                print("")
                row_cells = []

                for column in range(config.column_amount):
                    # posição no modelo x a razão da escala + (tamanho da celula e o espaço entre elas * coluna ou linha)
                    x1 = (getattr(config,f'grouping_{i}_x1') * x_axis_reason) + ((cell_size_x_px + x_space_between_cells) * column)
                    y1 = (getattr(config,f'grouping_{i}_y1') * y_axis_reason) + ((cell_size_y_px + y_space_between_cells) * row)
                    x1, y1 = handler.get_correct_positions(x1, y1)
                    cell = image[int(y1):int(y1+cell_size_y_px), int(x1):int(x1+cell_size_x_px)]
                    mean_color = cv2.mean(cell)[0]

                    if mean_color > highest_mean_color:
                        highest_mean_color = mean_color
                    elif mean_color < lowest_mean_color:
                        lowest_mean_color = mean_color

                    row_cells.append(mean_color)
                    print(str(int(mean_color))+"  ", end="")
                    cv2.rectangle(
                        image, (int(x1), int(y1)), 
                        (int(x1+cell_size_x_px), int(y1+cell_size_y_px)),
                        (0, 255, 0), 2
                    )
                grouping.append(row_cells)  
            groupings.append(grouping)          
        
        if groupings:
            info = {}

            color_diff = highest_mean_color - lowest_mean_color

            # toma como base a celula mais preenchida como 100% e a menos preenchida como 0% e assim define o necessário
            max_filled_cell_color = int(lowest_mean_color + ((1 - config.fill_precentage_to_consider_filled/100)*color_diff))
            max_doubtful_cell_color = int(lowest_mean_color + ((1- config.fill_precentage_to_consider_doubtful/100)*color_diff))
            
            cureent_question = 0
            for grouping in groupings:
                for row in grouping:
                    cureent_question += 1
                    ws_cell = self.ws[f"{ALPHABET[cureent_question]}{ws_row_index+1}"]
                    ws_cell.value = ''
                    for col_idx, col in enumerate(row):
                        if col <= max_filled_cell_color:
                            ws_cell.value = ws_cell.value + f' {ALPHABET[col_idx]}'
                        elif col <= max_doubtful_cell_color:
                            if cureent_question in info:
                                info[cureent_question].append(ALPHABET[col_idx]) 
                            else:
                                info[cureent_question] = [ALPHABET[col_idx]]
            self.information[ws_row_index] = info

            handler.save_cropped_image(os.path.join(self.answers_folder, f'interpreted_p{ws_row_index}.jpeg'))
            
            # cv2.imshow("Image with ROI", image)
            # cv2.waitKey(0)
            # cv2.destroyAllWindows() 